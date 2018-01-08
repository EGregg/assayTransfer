import csv, openpyxl, re
from openpyxl import Workbook
#for an exported results csv
#does not like the degree symbol

first = '3CSF.csv'
second = '2CSF.csv'
third = '1CSF.csv'
non_decimal = re.compile(r'[^\d.]+')
# e.g. non_decimal.sub('', '12.34fe4e')

book = openpyxl.load_workbook("sample.xlsx")
sheet = book.active

book.save("sample.xlsx")

def ResultsCSV(myresults):
	count = 12
	count2 = 12
	count3 = 12
	count4 = 12
	with open(myresults) as csvfile:
		readCSV = csv.reader(csvfile,delimiter=',')
		for row in readCSV:
			if not row:
				pass
			else:
				#row[1] is where the name is, row[2] has the Ct, row[3] has teh detector
				try:
					#print (row[1] + " " + row[3])
					#print (row[0] + ",",row[1]+",",row[2]+",",row[3],row[4]+",",row[5]+",")
					if "1e1" in row[1].lower() and "fam" in row[3].lower():
						#print ("1e1 found " + row[2])
						while sheet.cell(row=count,column=5).value != None and count <= 47:
							count += 1
						#sheet.cell(row=count,column=5).value = float(non_decimal.sub('',row[2]))
						sheet.cell(row=count,column=5).value = float(row[2])		

					elif "1e2" in row[1].lower() and "fam" in row[3].lower():
						#print ("1e2 found " + row[2])
						while sheet.cell(row=count2,column=6).value != None and count2 <= 47:
							count2 += 1
						sheet.cell(row=count2,column=6).value = float(row[2])	

					elif "1e3" in row[1].lower() and "fam" in row[3].lower():
						#print ("1e3 found " + row[2])
						while sheet.cell(row=count3,column=7).value != None and count3 <= 47:
							count3 += 1
						sheet.cell(row=count3,column=7).value = float(row[2])	

					elif "1e4" in row[1].lower() and "fam" in row[3].lower():
						#print ("1e4 found " + row[2])
						while sheet.cell(row=count4,column=8).value != None and count4 <= 47:
							count4 += 1
						sheet.cell(row=count4,column=8).value = float(row[2]) 
				except (IndexError):
					pass

		book.save("sample.xlsx")
	print ("Finished processing %s" % (myresults))

ResultsCSV(first)
ResultsCSV(second)
ResultsCSV(third)
		
'''name = row[1].split()
for temp in name:
print (temp)'''
