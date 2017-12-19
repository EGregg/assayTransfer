import csv, openpyxl
from openpyxl import Workbook
#https://www.reddit.com/r/learnpython/comments/63lvf7/really_confused_openpyxl_keeps_overwriting_my/
#the issue now is that it wants to overwrite the entire worksheet each time

book = openpyxl.load_workbook("sample.xlsx")
sheet = book.active
firstCol = 16
secondCol =16
thirdCol = 16
fourthCol = 16
count = 16

book.save("sample.xlsx")


with open('datacsv.csv') as csvfile:
	readCSV = csv.reader(csvfile,delimiter=',')
	for row in readCSV:
		if row == "":
			pass
		else:
			#print (row[0] + ",",row[1]+",",row[2])
			#row[1] is where the name is, row[2] has the Ct
			if "1e1" in row[1]:
				print ("1e1 found " + row[2])
				while sheet.cell(row=count,column=7).value != None:
					count += 1
				sheet.cell(row=count,column=7).value = row[2]			
				'''if sheet.cell(row=count,column=1).value != None:
					sheet.cell(row=count,column=1).value = row[2]					
					count += 1
				else:
					count += 1'''

			elif "1e2" in row[1]:
				print ("1e2 found " + row[2])
				sheet['F'+str(secondCol)] = row[2]
				secondCol = int(secondCol) + 1
				#book.save("sample.xlsx")

			elif "1e3" in row[1]:
				print ("1e3 found " + row[2])
				sheet['G'+str(thirdCol)] = row[2]
				thirdCol = int(thirdCol) + 1
				#book.save("sample.xlsx")

			elif "1e4" in row[1]:
				print ("1e4 found " + row[2])
				sheet['H'+str(fourthCol)] = row[2]
				fourthCol = int(fourthCol) + 1
				#book.save("sample.xlsx")

	book.save("sample.xlsx")


			
'''name = row[1].split()
for temp in name:
print (temp)'''
