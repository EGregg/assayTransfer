import csv, openpyxl, re
from openpyxl import Workbook
#for an exported results csv
#does not like the degree symbol
#for cleanupFile() -- UnicodeDecodeError: 'utf-8' codec can't decode byte 0xb0 in position 489: invalid start byte

firstList = []
first = 'first.csv'
second = 'second.csv'
third = 'third.csv'
non_decimal = re.compile(r'[^\d.]+')
# e.g. non_decimal.sub('', '12.34fe4e')

book = openpyxl.load_workbook("stabilityData.xlsx")
sheet = book.active
#sheet = book["FMD"]

book.save("stabilityData.xlsx")

#for myresults enter either first, second, or third where first, second, or third is the name of the file to be analyzed
def checker(myresults):
	with open(myresults) as csvfile:
		readCSV = csv.reader(csvfile,delimiter=',')
		for row in readCSV:
			if not row:
				pass
			else:
				print ("%s is row[1], %s is row[2], %s is row[3]" % (row[1],row[2],row[3]))	

#https://stackoverflow.com/questions/29725932/deleting-rows-with-python-in-a-csv-file
#takes a file and deletes the first 25 rows and spits out the edited cleanedup version
def cleanupFile(myresults):
	beginningRow = 25
	#changed the encoding to mac_roman because of the 'µ' symbol
	input1 = open(myresults,'rt',encoding='mac_roman')
	output1 = open('first_clean.csv','wt',encoding='mac_roman')	
	writer = csv.writer(output1)
	for row in csv.reader(input1):
		writer.writerow(row)	
	#input.close()
	#output.close()
	

#for myresults enter either first, second, or third where first, second, or third is the name of the file to be analyzed
def ResultsCSV(myresults):
	count = 12
	count2 = 12
	count3 = 12
	count4 = 12
	with open(myresults) as csvfile:
		readCSV = csv.reader(csvfile,delimiter=',')
		for row in readCSV:
			if not row:
				pass
			else:
				 #row[1] is where the name is, row[2] has the Ct, row[3] has teh detector
				if "FMD 1" in row[1]:
					sheet = book["FMD IVT 1"]
					print (row[1])				
				elif "fmd 2" in row[1].lower():
					sheet = book["FMD IVT 2"]
					print (row[1])
				try:
					#print (row[1] + " " + row[3])
					#print (row[0] + ",",row[1]+",",row[2]+",",row[3],row[4]+",",row[5]+",")
					if "1e1" in row[1].lower() and "fam" in row[3].lower():
						#print ("1e1 found " + row[2])
						while sheet.cell(row=count,column=3).value != None and count <= 24:
							count += 1
						#sheet.cell(row=count,column=5).value = float(non_decimal.sub('',row[2]))
						sheet.cell(row=count,column=3).value = float(row[2])		

					elif "1e2" in row[1].lower() and "fam" in row[3].lower():
						#print ("1e2 found " + row[2])
						while sheet.cell(row=count2,column=4).value != None and count2 <= 24:
							count2 += 1
						sheet.cell(row=count2,column=4).value = float(row[2])	

					elif "1e3" in row[1].lower() and "fam" in row[3].lower():
						#print ("1e3 found " + row[2])
						while sheet.cell(row=count3,column=5).value != None and count3 <= 24:
							count3 += 1
						sheet.cell(row=count3,column=5).value = float(row[2])	

					elif "1e4" in row[1].lower() and "fam" in row[3].lower():
						#print ("1e4 found " + row[2])
						while sheet.cell(row=count4,column=6).value != None and count4 <= 24:
							count4 += 1
						sheet.cell(row=count4,column=6).value = float(row[2]) 
				except (IndexError, ValueError):
					pass

		book.save("stabilityData.xlsx")
	print ("Finished processing %s" % (myresults))

def ResultsCSVcy5(myresults):
	column = 14	
	count1 = 12
	count2 = 12
	count3 = 12
	count4 = 12
	with open(myresults) as csvfile:
		readCSV = csv.reader(csvfile,delimiter=',')
		for row in readCSV:
			if not row:
				pass
			else:
				 #row[1] is where the name is, row[2] has the Ct, row[3] has the detector
				if "FMD 1" in row[1]:
					sheet = book["FMD IVT 1"]
					print ("%s is %s" % (row[1],row[3]))				
				elif "fmd 2" in row[1].lower() and "cy5" in row[3].lower():
					sheet = book["FMD IVT 2"]
					print ("%s is %s" % (row[1],row[2]))
				try:
					#print (row[1] + " " + row[3])
					#print (row[0] + ",",row[1]+",",row[2]+",",row[3],row[4]+",",row[5]+",")
					if "1e1" in row[1].lower() and "cy5" in row[3].lower():
						#print ("1e1 found " + row[2])
						while sheet.cell(row=count1,column=column).value != None and count1 <= 24:
							count1 += 1
						#sheet.cell(row=count,column=5).value = float(non_decimal.sub('',row[2]))
						sheet.cell(row=count1,column=column).value = float(row[2])		

					elif "1e2" in row[1].lower() and "cy5" in row[3].lower():
						#print ("1e2 found " + row[2])
						while sheet.cell(row=count2,column=column+1).value != None and count2 <= 24:
							count2 += 1
						sheet.cell(row=count2,column=column+1).value = float(row[2])	

					elif "1e3" in row[1].lower() and "cy5" in row[3].lower():
						#print ("1e3 found " + row[2])
						while sheet.cell(row=count3,column=column+2).value != None and count3 <= 24:
							count3 += 1
						sheet.cell(row=count3,column=column+2).value = float(row[2])	

					elif "1e4" in row[1].lower() and "cy5" in row[3].lower():
						#print ("1e4 found " + row[2])
						while sheet.cell(row=count4,column=column+3).value != None and count4 <= 24:
							count4 += 1
						sheet.cell(row=count4,column=column+3).value = float(row[2]) 
				except (IndexError, ValueError):
					pass

		book.save("stabilityData.xlsx")
	print ("Finished processing %s" % (myresults))



cleanupFile(first)

'''
ResultsCSVcy5(first)
ResultsCSV(second)
ResultsCSV(third)
'''

'''
checker(first)
'''

'''		
name = row[1].split()
for temp in name:
print (temp)'''
